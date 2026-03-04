# app.py (FULL) — Installable-safe version:
# - DB + settings stored in %APPDATA%\VARDA Control Center
# - Bundled resource loading for default PNGs (PyInstaller)
# - Works both in dev and after install

import sys
import re
import os
import time
import io
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import serial
import serial.tools.list_ports
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QFrame, QLabel, QPushButton, QLineEdit, QMessageBox,
    QComboBox, QFileDialog, QFormLayout, QGroupBox,
    QSlider, QDoubleSpinBox, QCheckBox, QInputDialog
)
from PySide6.QtGui import QPixmap
from PySide6.QtCore import Qt

from PIL import Image, ImageDraw, ImageFont
import qrcode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# Your db module functions
from db import (
    init_db, register_device,
    list_stores, create_store, assign_device_to_store,
    export_devices_rows
)

MODEL = "VR01"
MAC_RE = re.compile(r"^([0-9A-Fa-f]{2}:){5}[0-9A-Fa-f]{2}$")

# CHANGE THIS to your secret password
EDIT_LABEL_PASSWORD = "1234"

# Colors sampled from your example
BG_HEX = "#4781AA"
GOLD_HEX = "#D0B150"
WHITE = (255, 255, 255, 255)
BLACK = (0, 0, 0, 255)

PREVIEW_SIZE_PX = 420
LABEL_EXPORT_SIZE = 1200

# CP2102 VID/PID
CP2102_VID = 0x10C4
CP2102_PID = 0xEA60


# ----------------- INSTALLABLE PATHS -----------------

def app_data_dir() -> str:
    """
    Writable folder for DB + settings.
    Example: C:\\Users\\<User>\\AppData\\Roaming\\VARDA Control Center
    """
    base = os.getenv("APPDATA") or str(Path.home())
    p = Path(base) / "VARDA Control Center"
    p.mkdir(parents=True, exist_ok=True)
    return str(p)


def resource_path(relative: str) -> str:
    """
    Path to bundled resources under PyInstaller (sys._MEIPASS),
    or local folder during development.
    """
    base = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base, relative)


# ----------------- helpers -----------------

def hex_to_rgba(h: str, a=255):
    h = h.lstrip("#")
    return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16), a)


def _load_font(size: int):
    for path in [
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/SEGOEUI.TTF",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/ARIAL.TTF",
    ]:
        try:
            return ImageFont.truetype(path, size=size)
        except Exception:
            pass
    return ImageFont.load_default()


def _make_qr(data: str, size_px: int) -> Image.Image:
    qr = qrcode.QRCode(border=0, box_size=10)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    return img.resize((size_px, size_px), Image.LANCZOS)


def _rounded_mask(size: int, radius: int) -> Image.Image:
    m = Image.new("L", (size, size), 0)
    d = ImageDraw.Draw(m)
    d.rounded_rectangle([0, 0, size, size], radius=radius, fill=255)
    return m


def _load_png(path: str) -> Image.Image | None:
    if not path:
        return None
    if not os.path.exists(path):
        return None
    try:
        return Image.open(path).convert("RGBA")
    except Exception:
        return None


def _clamp(v: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, v))


def _safe_parse_datetime_to_excel(s: str):
    """
    Return a timezone-naive datetime for Excel.
    If DB timestamps are UTC (SQLite CURRENT_TIMESTAMP), convert to local PC time.
    """
    if not s:
        return s

    local_tz = datetime.now().astimezone().tzinfo or timezone.utc

    # Common SQLite format: "YYYY-MM-DD HH:MM:SS" (usually UTC)
    try:
        dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        dt = dt.replace(tzinfo=timezone.utc).astimezone(local_tz).replace(tzinfo=None)
        return dt
    except Exception:
        pass

    # ISO (may include timezone)
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt = dt.astimezone(local_tz).replace(tzinfo=None)
        return dt
    except Exception:
        return s


# ----------------- DB admin helpers (direct SQLite + schema-safe) -----------------

def _db_path() -> str:
    # Install-safe DB location (AppData)
    return os.path.join(app_data_dir(), "devices.db")


def _table_columns(cur: sqlite3.Cursor, table: str) -> list[str]:
    try:
        cur.execute(f"PRAGMA table_info({table});")
        return [r[1] for r in cur.fetchall()]  # name at index 1
    except Exception:
        return []


def _pick_first(cols: list[str], candidates: list[str]) -> str | None:
    s = set(cols)
    for c in candidates:
        if c in s:
            return c
    return None


def _table_exists(cur: sqlite3.Cursor, table: str) -> bool:
    try:
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?;", (table,))
        return cur.fetchone() is not None
    except Exception:
        return False


def clear_devices_only():
    """
    Erase device data ONLY:
    - deletes from device_store
    - deletes from devices
    - keeps stores table untouched
    """
    path = _db_path()
    con = sqlite3.connect(path)
    try:
        cur = con.cursor()
        cur.execute("PRAGMA foreign_keys=OFF;")

        if _table_exists(cur, "device_store"):
            cur.execute("DELETE FROM device_store;")

        if _table_exists(cur, "devices"):
            cur.execute("DELETE FROM devices;")

        try:
            cur.execute("DELETE FROM sqlite_sequence WHERE name IN ('devices','device_store');")
        except Exception:
            pass

        con.commit()
    finally:
        con.close()


def delete_one_device_by_serial(serial_value: str) -> int:
    """
    Delete a single device by serial.
    Removes mappings in device_store (by serial or device_id if present).
    Returns number of rows deleted from devices (0 or 1).
    """
    serial_value = (serial_value or "").strip()
    if not serial_value:
        return 0

    path = _db_path()
    con = sqlite3.connect(path)
    try:
        cur = con.cursor()
        cur.execute("PRAGMA foreign_keys=OFF;")

        if not _table_exists(cur, "devices"):
            return 0

        dev_cols = _table_columns(cur, "devices")
        serial_col = _pick_first(dev_cols, ["serial", "device_serial", "sn"])
        id_col = _pick_first(dev_cols, ["id", "device_id"])

        if not serial_col:
            return 0

        device_id = None
        if id_col:
            try:
                cur.execute(f"SELECT {id_col} FROM devices WHERE {serial_col}=?;", (serial_value,))
                row = cur.fetchone()
                device_id = row[0] if row else None
            except Exception:
                device_id = None

        if _table_exists(cur, "device_store"):
            ds_cols = _table_columns(cur, "device_store")
            ds_device_id_col = _pick_first(ds_cols, ["device_id", "deviceId"])
            ds_serial_col = _pick_first(ds_cols, ["device_serial", "serial", "sn"])

            if device_id is not None and ds_device_id_col:
                try:
                    cur.execute(f"DELETE FROM device_store WHERE {ds_device_id_col}=?;", (device_id,))
                except Exception:
                    pass
            elif ds_serial_col:
                try:
                    cur.execute(f"DELETE FROM device_store WHERE {ds_serial_col}=?;", (serial_value,))
                except Exception:
                    pass

        deleted = 0
        try:
            cur.execute(f"DELETE FROM devices WHERE {serial_col}=?;", (serial_value,))
            deleted = cur.rowcount or 0
        except Exception:
            deleted = 0

        con.commit()
        return deleted
    finally:
        con.close()


def delete_one_store_by_id(store_id: int) -> int:
    """
    Delete a single store by its ID.
    Removes mappings in device_store referencing that store.
    Devices stay untouched.
    Returns number of store rows deleted (0 or 1).
    """
    path = _db_path()
    con = sqlite3.connect(path)
    try:
        cur = con.cursor()
        cur.execute("PRAGMA foreign_keys=OFF;")

        if not _table_exists(cur, "stores"):
            return 0

        stores_cols = _table_columns(cur, "stores")
        stores_id_col = _pick_first(stores_cols, ["id", "store_id"])
        if not stores_id_col:
            return 0

        if _table_exists(cur, "device_store"):
            ds_cols = _table_columns(cur, "device_store")
            ds_store_id_col = _pick_first(ds_cols, ["store_id", "storeId"])
            if ds_store_id_col:
                try:
                    cur.execute(f"DELETE FROM device_store WHERE {ds_store_id_col}=?;", (int(store_id),))
                except Exception:
                    pass

        deleted = 0
        try:
            cur.execute(f"DELETE FROM stores WHERE {stores_id_col}=?;", (int(store_id),))
            deleted = cur.rowcount or 0
        except Exception:
            deleted = 0

        con.commit()
        return deleted
    finally:
        con.close()


# ----------------- USB DETECTION (SAFE + FAST) -----------------

def list_com_port_infos():
    return list(serial.tools.list_ports.comports())


def is_cp2102(port_info) -> bool:
    vid = getattr(port_info, "vid", None)
    pid = getattr(port_info, "pid", None)
    return (vid == CP2102_VID and pid == CP2102_PID)


def try_read_varda_marker_fast(port: str) -> bool:
    try:
        with serial.Serial(port, baudrate=115200, timeout=0.20) as ser:
            time.sleep(0.08)
            for _ in range(8):
                line = ser.readline().decode(errors="ignore").strip()
                if line.startswith("VARDA_DEVICE=") or line.startswith("MAC="):
                    return True
    except Exception:
        return False
    return False


def auto_detect_esp32_port():
    ports = list_com_port_infos()
    cp_ports = [p for p in ports if is_cp2102(p)]
    if not cp_ports:
        return None

    for p in cp_ports:
        if try_read_varda_marker_fast(p.device):
            return p.device

    return cp_ports[0].device


def read_mac_from_port_fast(port: str) -> str | None:
    try:
        with serial.Serial(port, baudrate=115200, timeout=0.30) as ser:
            time.sleep(0.10)
            for _ in range(15):
                line = ser.readline().decode(errors="ignore").strip()
                if line.startswith("MAC="):
                    return line.split("=", 1)[1].strip().upper()
    except Exception:
        return None
    return None


# ----------------- Label renderer (from scratch) -----------------

def render_label(
    size: int,
    subtitle: str,
    tiny_note: str,
    serial_value: str,
    mac_value: str,
    user_value: str,
    password_value: str,
    input_value: str,
    top_right_text: str,
    bottom_center_text: str,
    qr1_data: str,
    qr2_data: str,
    qr1_caption: str,
    qr2_caption: str,
    img_logo_path: str,
    img_icon1_path: str,
    img_icon2_path: str,
    img_icon3_path: str,
    text_scale: float,
    logo_scale: float,
    qr_scale: float,
    icon_scale: float,

    logo_x: float,
    logo_y: float,

    subtitle_y: float,
    tiny_note_y: float,
    info_x: float,
    info_y: float,
    top_right_x: float,
    top_right_y: float,
    bottom_center_y: float,
) -> Image.Image:
    bg = hex_to_rgba(BG_HEX)
    img = Image.new("RGBA", (size, size), bg)
    draw = ImageDraw.Draw(img)

    scale = size
    radius = int(scale * 0.11)
    mask = _rounded_mask(size, radius)
    img.putalpha(mask)

    f_sub = _load_font(max(10, int(scale * 0.040 * text_scale)))
    f_tiny = _load_font(max(9, int(scale * 0.028 * text_scale)))
    f_text = _load_font(max(10, int(scale * 0.040 * text_scale)))
    f_small = _load_font(max(9, int(scale * 0.030 * text_scale)))

    # Logo (movable)
    logo = _load_png(img_logo_path)
    if logo:
        target_w = int(scale * 0.45 * logo_scale)
        target_h = int(scale * 0.14 * logo_scale)
        logo2 = logo.copy()
        logo2.thumbnail((target_w, target_h), Image.LANCZOS)

        cx = int(scale * float(logo_x))
        ly = int(scale * float(logo_y))
        lx = cx - (logo2.size[0] // 2)

        lx = _clamp(lx, 0, max(0, scale - logo2.size[0]))
        ly = _clamp(ly, 0, max(0, scale - logo2.size[1]))
        img.alpha_composite(logo2, (lx, ly))
    else:
        gold = hex_to_rgba(GOLD_HEX)
        f_title = _load_font(max(10, int(scale * 0.070 * text_scale)))
        txt = "Varda"
        tw = draw.textlength(txt, font=f_title)
        draw.text(((scale - tw) / 2, int(scale * 0.06)), txt, fill=gold, font=f_title)

    # Subtitle
    if subtitle:
        sub_y_px = int(scale * float(subtitle_y))
        sub_w = draw.textlength(subtitle, font=f_sub)
        draw.text(((scale - sub_w) / 2, sub_y_px), subtitle, fill=BLACK, font=f_sub)

    # Tiny note
    if tiny_note:
        ny_px = int(scale * float(tiny_note_y))
        nw = draw.textlength(tiny_note, font=f_tiny)
        draw.text(((scale - nw) / 2, ny_px), tiny_note, fill=BLACK, font=f_tiny)

    # Top-right small text
    if top_right_text:
        tx = int(scale * float(top_right_x))
        ty = int(scale * float(top_right_y))
        draw.text((tx, ty), top_right_text, fill=BLACK, font=f_tiny)

    # Info block
    x0 = int(scale * float(info_x))
    y0 = int(scale * float(info_y))
    line_h = int(scale * 0.075 * text_scale)

    lines = [
        f"S/N: {serial_value}",
        f"MAC: {mac_value}",
        "",
        f"User: {user_value}",
        f"Password: {password_value}",
        "",
        f"Input: {input_value}",
    ]

    y = y0
    for line in lines:
        if line == "":
            y += int(line_h * 0.5)
            continue
        draw.text((x0, y), line, fill=WHITE, font=f_text)
        y += line_h

    # QRs
    qr_size = int(scale * 0.20 * qr_scale)
    qr_y = int(scale * 0.69)
    left_qr_x = int(scale * 0.10)
    right_qr_x = int(scale * 0.70)

    if qr1_data:
        img.alpha_composite(_make_qr(qr1_data, qr_size), (left_qr_x, qr_y))
    if qr2_data:
        img.alpha_composite(_make_qr(qr2_data, qr_size), (right_qr_x, qr_y))

    cap_y = qr_y + qr_size + int(scale * 0.02)
    if qr1_caption:
        cap_w = draw.textlength(qr1_caption, font=f_small)
        draw.text((left_qr_x + (qr_size - cap_w) / 2, cap_y), qr1_caption, fill=WHITE, font=f_small)
    if qr2_caption:
        cap_w = draw.textlength(qr2_caption, font=f_small)
        draw.text((right_qr_x + (qr_size - cap_w) / 2, cap_y), qr2_caption, fill=WHITE, font=f_small)

    # Bottom center text
    if bottom_center_text:
        by = int(scale * float(bottom_center_y))
        bw = draw.textlength(bottom_center_text, font=f_tiny)
        draw.text(((scale - bw) / 2, by), bottom_center_text, fill=WHITE, font=f_tiny)

    # Bottom icons (3)
    icons = [_load_png(img_icon1_path), _load_png(img_icon2_path), _load_png(img_icon3_path)]
    icons = [ic for ic in icons if ic is not None]
    icon_size = int(scale * 0.07 * icon_scale)
    gap = int(scale * 0.06)
    icons_total_w = len(icons) * icon_size + (len(icons) - 1) * gap if icons else 0
    icons_y = int(scale * 0.90)

    ix = (scale - icons_total_w) // 2 if icons_total_w else 0
    for ic in icons:
        ic2 = ic.copy()
        ic2.thumbnail((icon_size, icon_size), Image.LANCZOS)
        img.alpha_composite(ic2, (ix + (icon_size - ic2.size[0]) // 2, icons_y))
        ix += icon_size + gap

    return img


def label_to_qpix(img: Image.Image, preview_size: int) -> QPixmap:
    preview = img.resize((preview_size, preview_size), Image.LANCZOS)
    buf = io.BytesIO()
    preview.save(buf, format="PNG")
    qpix = QPixmap()
    qpix.loadFromData(buf.getvalue(), "PNG")
    return qpix


# ----------------- Excel export (pro level) -----------------

def export_devices_to_excel_pro(path: str):
    headers = ["serial", "mac", "model", "batch", "store_name", "admin_password", "created_at"]
    rows = export_devices_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    last_col = len(headers)
    last_col_letter = ws.cell(row=1, column=last_col).column_letter

    ws["A1"] = "VARDA Control Center — Production Log"
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A2"] = "Exported at (PC time):"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "Total devices:"
    ws["B3"] = len(rows)

    header_row = 5
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h.upper())

    start_data_row = header_row + 1
    for r_i, r in enumerate(rows, start=start_data_row):
        r = list(r)
        if len(r) >= 7:
            r[6] = _safe_parse_datetime_to_excel(str(r[6]) if r[6] is not None else "")
            if isinstance(r[6], datetime) and r[6].tzinfo is not None:
                r[6] = r[6].replace(tzinfo=None)

        for c_i, val in enumerate(r, start=1):
            ws.cell(row=r_i, column=c_i, value=val)

    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="0F172A")
    meta_fill = PatternFill("solid", fgColor="111827")
    meta_font = Font(bold=True, color="E5E7EB")
    meta_val_font = Font(color="E5E7EB")

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    body_align = Alignment(vertical="center")

    thin = Side(style="thin", color="243046")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, 4):
        ws[f"A{row}"].font = meta_font
        ws[f"B{row}"].font = meta_val_font
        ws[f"A{row}"].fill = meta_fill
        ws[f"B{row}"].fill = meta_fill
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border

    for c in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=start_data_row, max_row=max_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.alignment = body_align
            cell.border = border

    created_at_col = 7
    for cell in ws.iter_rows(min_row=start_data_row, max_row=max_row, min_col=created_at_col, max_col=created_at_col):
        c = cell[0]
        if isinstance(c.value, datetime):
            c.number_format = "yyyy-mm-dd hh:mm:ss"

    ws.freeze_panes = ws.cell(row=start_data_row, column=1)

    table_ref = f"A{header_row}:{last_col_letter}{max_row}"
    tab = Table(displayName="DevicesTable", ref=table_ref)
    tab_style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = tab_style
    ws.add_table(tab)

    widths = [len(h) for h in headers]
    for r in ws.iter_rows(min_row=header_row, max_row=max_row, min_col=1, max_col=last_col):
        for i, cell in enumerate(r):
            v = "" if cell.value is None else str(cell.value)
            widths[i] = min(max(widths[i], len(v)), 60)

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = max(12, min(w + 2, 60))

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[header_row].height = 22
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110

    folder = os.path.dirname(path)
    if folder:
        os.makedirs(folder, exist_ok=True)

    wb.save(path)


# ----------------- MAIN WINDOW -----------------

class MainWindow(QMainWindow):
    SETTINGS_FILE = os.path.join(app_data_dir(), "settings.json")

    def __init__(self):
        super().__init__()
        self.setWindowTitle("VARDA Control Center")
        from PySide6.QtGui import QIcon
        self.setWindowIcon(QIcon(resource_path("varda.ico")))
        self.resize(1380, 800)

        # Default bundled images
        self.img_logo_path = resource_path("img_logo.png")
        self.img_icon1_path = resource_path("img_icon1.png")
        self.img_icon2_path = resource_path("img_icon2.png")
        self.img_icon3_path = resource_path("img_icon3.png")

        self.excel_export_path = None

        root = QWidget()
        root_layout = QHBoxLayout(root)
        root_layout.setContentsMargins(12, 12, 12, 12)
        root_layout.setSpacing(12)

        # Sidebar
        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(14, 14, 14, 14)
        sidebar_layout.setSpacing(10)

        brand = QLabel("VARDA")
        brand.setObjectName("brand")

        btn_export_excel = QPushButton("Export Excel (.xlsx)")
        btn_export_excel.setObjectName("navButton")
        btn_export_excel.clicked.connect(self.on_export_excel)

        btn_export_label = QPushButton("Export Label PNG")
        btn_export_label.setObjectName("navButton")
        btn_export_label.clicked.connect(self.on_export_label_png)

        btn_delete_device = QPushButton("Delete one device…")
        btn_delete_device.setObjectName("navButton")
        btn_delete_device.clicked.connect(self.on_delete_one_device)

        btn_delete_store = QPushButton("Delete one store…")
        btn_delete_store.setObjectName("navButton")
        btn_delete_store.clicked.connect(self.on_delete_one_store)

        btn_clear_devices = QPushButton("Clear devices data…")
        btn_clear_devices.setObjectName("dangerButton")
        btn_clear_devices.clicked.connect(self.on_clear_devices_only)

        sidebar_layout.addWidget(brand)
        sidebar_layout.addSpacing(10)
        sidebar_layout.addWidget(btn_export_excel)
        sidebar_layout.addWidget(btn_export_label)
        sidebar_layout.addSpacing(10)
        sidebar_layout.addWidget(btn_delete_device)
        sidebar_layout.addWidget(btn_delete_store)
        sidebar_layout.addWidget(btn_clear_devices)

        sidebar_layout.addSpacing(12)

        grp_img = QGroupBox("Images (4)")
        grp_img.setObjectName("group")
        img_layout = QVBoxLayout(grp_img)
        img_layout.setSpacing(8)

        self.lbl_logo = QLabel(f"Logo: {self.img_logo_path}")
        self.lbl_i1 = QLabel(f"Icon 1: {self.img_icon1_path}")
        self.lbl_i2 = QLabel(f"Icon 2: {self.img_icon2_path}")
        self.lbl_i3 = QLabel(f"Icon 3: {self.img_icon3_path}")

        btn_logo = QPushButton("Choose Logo PNG"); btn_logo.setObjectName("navButton")
        btn_logo.clicked.connect(lambda: self.pick_image("logo"))

        btn_i1 = QPushButton("Choose Icon 1 PNG"); btn_i1.setObjectName("navButton")
        btn_i1.clicked.connect(lambda: self.pick_image("i1"))

        btn_i2 = QPushButton("Choose Icon 2 PNG"); btn_i2.setObjectName("navButton")
        btn_i2.clicked.connect(lambda: self.pick_image("i2"))

        btn_i3 = QPushButton("Choose Icon 3 PNG"); btn_i3.setObjectName("navButton")
        btn_i3.clicked.connect(lambda: self.pick_image("i3"))

        img_layout.addWidget(self.lbl_logo); img_layout.addWidget(btn_logo)
        img_layout.addWidget(self.lbl_i1); img_layout.addWidget(btn_i1)
        img_layout.addWidget(self.lbl_i2); img_layout.addWidget(btn_i2)
        img_layout.addWidget(self.lbl_i3); img_layout.addWidget(btn_i3)

        sidebar_layout.addWidget(grp_img)

        grp_pos = QGroupBox("Text positions")
        grp_pos.setObjectName("group")
        pos_form = QFormLayout(grp_pos)

        def make_pos(default):
            sb = QDoubleSpinBox()
            sb.setRange(0.0, 1.0)
            sb.setSingleStep(0.01)
            sb.setDecimals(2)
            sb.setValue(default)
            sb.valueChanged.connect(self.update_preview)
            return sb

        self.pos_subtitle_y = make_pos(0.23)
        self.pos_tiny_note_y = make_pos(0.285)
        self.pos_info_x = make_pos(0.10)
        self.pos_info_y = make_pos(0.32)
        self.pos_top_right_x = make_pos(0.70)
        self.pos_top_right_y = make_pos(0.06)
        self.pos_bottom_center_y = make_pos(0.86)

        pos_form.addRow("Subtitle Y", self.pos_subtitle_y)
        pos_form.addRow("Tiny note Y", self.pos_tiny_note_y)
        pos_form.addRow("Info X", self.pos_info_x)
        pos_form.addRow("Info Y", self.pos_info_y)
        pos_form.addRow("Top-right X", self.pos_top_right_x)
        pos_form.addRow("Top-right Y", self.pos_top_right_y)
        pos_form.addRow("Bottom text Y", self.pos_bottom_center_y)

        sidebar_layout.addWidget(grp_pos)
        sidebar_layout.addStretch(1)

        # Content
        content = QFrame()
        content.setObjectName("content")
        content_layout = QHBoxLayout(content)
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(18)

        # LEFT
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setSpacing(12)

        title = QLabel("Register device")
        title.setObjectName("title")

        self.port_combo = QComboBox()
        self.port_combo.setObjectName("input")

        btn_refresh_ports = QPushButton("Refresh Ports")
        btn_refresh_ports.setObjectName("navButton")
        btn_refresh_ports.clicked.connect(self.refresh_ports)

        btn_auto_detect = QPushButton("Auto-detect ESP32")
        btn_auto_detect.setObjectName("navButton")
        btn_auto_detect.clicked.connect(self.on_auto_detect)

        btn_read_mac = QPushButton("Read MAC from USB")
        btn_read_mac.setObjectName("readMacBtn")
        btn_read_mac.clicked.connect(self.on_read_mac)

        self.store_combo = QComboBox()
        self.store_combo.setObjectName("input")

        self.new_store = QLineEdit()
        self.new_store.setPlaceholderText("Add new store (optional)")
        self.new_store.setObjectName("input")

        btn_add_store = QPushButton("Add store")
        btn_add_store.setObjectName("navButton")
        btn_add_store.clicked.connect(self.on_add_store)

        grp = QGroupBox("Label fields")
        grp.setObjectName("group")
        grp_layout = QVBoxLayout(grp)
        grp_layout.setSpacing(10)

        self.cb_edit_label = QCheckBox("Edit label fields (locked)")
        self.cb_edit_label.setChecked(False)
        self.cb_edit_label.toggled.connect(self.on_toggle_edit_label)
        grp_layout.addWidget(self.cb_edit_label)

        form = QFormLayout()
        grp_layout.addLayout(form)

        self.f_subtitle = QLineEdit("NFC digital receipts"); self.f_subtitle.setObjectName("input")
        self.f_tiny_note = QLineEdit(""); self.f_tiny_note.setPlaceholderText("Optional note under subtitle"); self.f_tiny_note.setObjectName("input")

        self.f_serial = QLineEdit("VR01-YYMM-0001"); self.f_serial.setObjectName("input")
        self.f_mac = QLineEdit("AA:BB:CC:DD:EE:FF"); self.f_mac.setObjectName("input")

        self.f_user = QLineEdit("Admin123"); self.f_user.setObjectName("input")
        self.f_password = QLineEdit("12345678"); self.f_password.setObjectName("input")
        self.f_input = QLineEdit("5V ⎓ 1A"); self.f_input.setObjectName("input")

        self.f_top_right = QLineEdit("VR01"); self.f_top_right.setObjectName("input")
        self.f_bottom_center = QLineEdit("varda.app"); self.f_bottom_center.setObjectName("input")

        self.f_qr1 = QLineEdit("https://varda.example/setup-guide"); self.f_qr1.setObjectName("input")
        self.f_qr2 = QLineEdit("http://192.168.4.1"); self.f_qr2.setObjectName("input")
        self.f_qr1_cap = QLineEdit("Setup guide"); self.f_qr1_cap.setObjectName("input")
        self.f_qr2_cap = QLineEdit("Wi-Fi Setup"); self.f_qr2_cap.setObjectName("input")

        form.addRow("Subtitle", self.f_subtitle)
        form.addRow("Tiny note", self.f_tiny_note)
        form.addRow("S/N", self.f_serial)
        form.addRow("MAC", self.f_mac)
        form.addRow("User", self.f_user)
        form.addRow("Password", self.f_password)
        form.addRow("Input", self.f_input)
        form.addRow("Top-right text", self.f_top_right)
        form.addRow("Bottom-center text", self.f_bottom_center)

        btn_save = QPushButton("Generate + Save (DB)")
        btn_save.setObjectName("saveBtn")
        btn_save.clicked.connect(self.on_save)

        self.status1 = QLabel("Serial: —"); self.status1.setObjectName("muted")
        self.status2 = QLabel("Admin password: —"); self.status2.setObjectName("muted")

        left_layout.addWidget(title)

        top_row = QHBoxLayout()
        top_row.setSpacing(14)

        com_col = QVBoxLayout()
        com_col.setSpacing(8)
        com_col.addWidget(QLabel("ESP32 USB Port (COM)"))
        com_col.addWidget(self.port_combo)

        com_btns = QHBoxLayout()
        com_btns.setSpacing(8)
        com_btns.addWidget(btn_refresh_ports)
        com_btns.addWidget(btn_auto_detect)
        com_col.addLayout(com_btns)
        com_col.addWidget(btn_read_mac)

        store_col = QVBoxLayout()
        store_col.setSpacing(8)
        store_col.addWidget(QLabel("Store"))
        store_col.addWidget(self.store_combo)
        store_col.addWidget(self.new_store)
        store_col.addWidget(btn_add_store)

        top_row.addLayout(com_col, 1)
        top_row.addLayout(store_col, 1)

        left_layout.addLayout(top_row)
        left_layout.addWidget(grp)
        left_layout.addWidget(btn_save)
        left_layout.addWidget(self.status1)
        left_layout.addWidget(self.status2)
        left_layout.addStretch(1)

        # RIGHT
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setSpacing(10)

        preview_title = QLabel("Label preview")
        preview_title.setObjectName("titleSmall")

        self.label_preview = QLabel()
        self.label_preview.setObjectName("preview")
        self.label_preview.setFixedSize(PREVIEW_SIZE_PX, PREVIEW_SIZE_PX)

        grp_set = QGroupBox("Label settings (live)")
        grp_set.setObjectName("group")
        set_form = QFormLayout(grp_set)

        self.s_text = QSlider(Qt.Horizontal); self.s_text.setMinimum(70); self.s_text.setMaximum(140); self.s_text.setValue(100)
        self.s_logo = QSlider(Qt.Horizontal); self.s_logo.setMinimum(70); self.s_logo.setMaximum(140); self.s_logo.setValue(100)
        self.s_qr = QSlider(Qt.Horizontal); self.s_qr.setMinimum(70); self.s_qr.setMaximum(140); self.s_qr.setValue(100)
        self.s_icon = QSlider(Qt.Horizontal); self.s_icon.setMinimum(70); self.s_icon.setMaximum(140); self.s_icon.setValue(100)

        set_form.addRow("Text size", self.s_text)
        set_form.addRow("Logo size", self.s_logo)
        set_form.addRow("QR size", self.s_qr)
        set_form.addRow("Icons size", self.s_icon)

        self.pos_logo_x = QDoubleSpinBox()
        self.pos_logo_x.setRange(0.0, 1.0)
        self.pos_logo_x.setSingleStep(0.01)
        self.pos_logo_x.setDecimals(2)
        self.pos_logo_x.setValue(0.50)

        self.pos_logo_y = QDoubleSpinBox()
        self.pos_logo_y.setRange(0.0, 1.0)
        self.pos_logo_y.setSingleStep(0.01)
        self.pos_logo_y.setDecimals(2)
        self.pos_logo_y.setValue(0.06)

        set_form.addRow("Logo X (0→1)", self.pos_logo_x)
        set_form.addRow("Logo Y (0→1)", self.pos_logo_y)

        set_form.addRow("QR1 Data", self.f_qr1)
        set_form.addRow("QR1 Caption", self.f_qr1_cap)
        set_form.addRow("QR2 Data", self.f_qr2)
        set_form.addRow("QR2 Caption", self.f_qr2_cap)

        right_layout.addWidget(preview_title)
        right_layout.addWidget(self.label_preview)
        right_layout.addWidget(grp_set)
        right_layout.addStretch(1)

        content_layout.addWidget(left, 2)
        content_layout.addWidget(right, 1)

        root_layout.addWidget(sidebar)
        root_layout.addWidget(content, 1)
        self.setCentralWidget(root)

        self.setStyleSheet("""
            QMainWindow { background: #0B1220; }

            QFrame#sidebar {
                background: #0F172A;
                border: 1px solid #1F2A3D;
                border-radius: 16px;
                min-width: 300px;
            }

            QFrame#content {
                background: #0F172A;
                border: 1px solid #1F2A3D;
                border-radius: 16px;
            }

            QLabel { color: #E5E7EB; font-size: 13px; }
            QLabel#brand { font-size: 20px; font-weight: 800; }
            QLabel#title { font-size: 18px; font-weight: 700; }
            QLabel#titleSmall { font-size: 15px; font-weight: 700; }
            QLabel#muted { color: #94A3B8; }

            QGroupBox#group {
                color: #E5E7EB;
                border: 1px solid #1F2A3D;
                border-radius: 12px;
                margin-top: 10px;
                padding: 10px;
            }
            QGroupBox#group::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 6px;
            }

            QLineEdit#input, QComboBox#input {
                background: #0B1220;
                border: 1px solid #1F2A3D;
                border-radius: 12px;
                padding: 10px;
                color: #E5E7EB;
            }

            QComboBox#input::drop-down {
                border: none;
                width: 30px;
                background: transparent;
            }
            QComboBox#input::down-arrow {
                image: none;
            }

            QPushButton#navButton {
                background: transparent;
                border: 1px solid #1F2A3D;
                border-radius: 12px;
                padding: 10px;
                color: #E5E7EB;
            }
            QPushButton#navButton:hover { background: #0B1220; }

            QPushButton#saveBtn, QPushButton#readMacBtn {
                background: #2563EB;
                border: none;
                border-radius: 12px;
                padding: 12px;
                color: white;
                font-weight: 700;
            }
            QPushButton#saveBtn:hover, QPushButton#readMacBtn:hover {
                background: #1D4ED8;
            }

            QPushButton#dangerButton {
                background: #7F1D1D;
                border: 1px solid #991B1B;
                border-radius: 12px;
                padding: 10px;
                color: #FEE2E2;
                font-weight: 700;
            }
            QPushButton#dangerButton:hover {
                background: #991B1B;
            }

            QLabel#preview {
                background: #0B1220;
                border: 1px solid #1F2A3D;
                border-radius: 16px;
            }
        """)

        for w in [
            self.f_subtitle, self.f_tiny_note,
            self.f_serial, self.f_mac,
            self.f_user, self.f_password,
            self.f_input, self.f_top_right, self.f_bottom_center,
            self.f_qr1, self.f_qr2, self.f_qr1_cap, self.f_qr2_cap
        ]:
            w.textChanged.connect(self.update_preview)

        for s in [self.s_text, self.s_logo, self.s_qr, self.s_icon]:
            s.valueChanged.connect(self.update_preview)

        self.pos_logo_x.valueChanged.connect(self.update_preview)
        self.pos_logo_y.valueChanged.connect(self.update_preview)

        self.refresh_ports()
        self.refresh_stores()
        self.load_settings_from_disk()

        self.set_label_fields_enabled(False)
        self.cb_edit_label.blockSignals(True)
        self.cb_edit_label.setChecked(False)
        self.cb_edit_label.blockSignals(False)
        self.cb_edit_label.setText("Edit label fields (locked)")

        self.update_preview()
        self.on_auto_detect()

    # -------- password-locked label editing --------

    def set_label_fields_enabled(self, enabled: bool):
        edits = [
            self.f_subtitle, self.f_tiny_note,
            self.f_serial, self.f_mac,
            self.f_user, self.f_password, self.f_input,
            self.f_top_right, self.f_bottom_center,
            self.f_qr1, self.f_qr2, self.f_qr1_cap, self.f_qr2_cap,
        ]
        for w in edits:
            w.setReadOnly(not enabled)

    def on_toggle_edit_label(self, checked: bool):
        if checked:
            self.cb_edit_label.blockSignals(True)
            self.cb_edit_label.setChecked(False)
            self.cb_edit_label.blockSignals(False)

            pwd, ok = QInputDialog.getText(
                self,
                "Unlock editing",
                "Enter password to edit label fields:",
                QLineEdit.Password
            )
            if not ok:
                self.set_label_fields_enabled(False)
                self.cb_edit_label.setText("Edit label fields (locked)")
                return

            if pwd != EDIT_LABEL_PASSWORD:
                QMessageBox.warning(self, "Wrong password", "Password is incorrect.")
                self.set_label_fields_enabled(False)
                self.cb_edit_label.setText("Edit label fields (locked)")
                return

            self.set_label_fields_enabled(True)
            self.cb_edit_label.blockSignals(True)
            self.cb_edit_label.setChecked(True)
            self.cb_edit_label.blockSignals(False)
            self.cb_edit_label.setText("Edit label fields (unlocked)")
            return

        self.set_label_fields_enabled(False)
        self.cb_edit_label.setText("Edit label fields (locked)")

    # -------- settings persistence --------

    def collect_settings(self) -> dict:
        return {
            "subtitle": self.f_subtitle.text(),
            "tiny_note": self.f_tiny_note.text(),
            "serial": self.f_serial.text(),
            "mac": self.f_mac.text(),
            "user": self.f_user.text(),
            "password": self.f_password.text(),
            "input": self.f_input.text(),
            "top_right": self.f_top_right.text(),
            "bottom_center": self.f_bottom_center.text(),

            "qr1": self.f_qr1.text(),
            "qr2": self.f_qr2.text(),
            "qr1_cap": self.f_qr1_cap.text(),
            "qr2_cap": self.f_qr2_cap.text(),

            "s_text": self.s_text.value(),
            "s_logo": self.s_logo.value(),
            "s_qr": self.s_qr.value(),
            "s_icon": self.s_icon.value(),

            "pos_logo_x": float(self.pos_logo_x.value()),
            "pos_logo_y": float(self.pos_logo_y.value()),

            "pos_subtitle_y": float(self.pos_subtitle_y.value()),
            "pos_tiny_note_y": float(self.pos_tiny_note_y.value()),
            "pos_info_x": float(self.pos_info_x.value()),
            "pos_info_y": float(self.pos_info_y.value()),
            "pos_top_right_x": float(self.pos_top_right_x.value()),
            "pos_top_right_y": float(self.pos_top_right_y.value()),
            "pos_bottom_center_y": float(self.pos_bottom_center_y.value()),

            "img_logo_path": self.img_logo_path,
            "img_icon1_path": self.img_icon1_path,
            "img_icon2_path": self.img_icon2_path,
            "img_icon3_path": self.img_icon3_path,

            "selected_store_id": self.store_combo.currentData(),
            "excel_export_path": self.excel_export_path,
        }

    def apply_settings(self, data: dict):
        self.f_subtitle.setText(data.get("subtitle", self.f_subtitle.text()))
        self.f_tiny_note.setText(data.get("tiny_note", self.f_tiny_note.text()))
        self.f_serial.setText(data.get("serial", self.f_serial.text()))
        self.f_mac.setText(data.get("mac", self.f_mac.text()))
        self.f_user.setText(data.get("user", self.f_user.text()))
        self.f_password.setText(data.get("password", self.f_password.text()))
        self.f_input.setText(data.get("input", self.f_input.text()))
        self.f_top_right.setText(data.get("top_right", self.f_top_right.text()))
        self.f_bottom_center.setText(data.get("bottom_center", self.f_bottom_center.text()))

        self.f_qr1.setText(data.get("qr1", self.f_qr1.text()))
        self.f_qr2.setText(data.get("qr2", self.f_qr2.text()))
        self.f_qr1_cap.setText(data.get("qr1_cap", self.f_qr1_cap.text()))
        self.f_qr2_cap.setText(data.get("qr2_cap", self.f_qr2_cap.text()))

        self.s_text.setValue(int(data.get("s_text", self.s_text.value())))
        self.s_logo.setValue(int(data.get("s_logo", self.s_logo.value())))
        self.s_qr.setValue(int(data.get("s_qr", self.s_qr.value())))
        self.s_icon.setValue(int(data.get("s_icon", self.s_icon.value())))

        self.pos_logo_x.setValue(float(data.get("pos_logo_x", self.pos_logo_x.value())))
        self.pos_logo_y.setValue(float(data.get("pos_logo_y", self.pos_logo_y.value())))

        self.pos_subtitle_y.setValue(float(data.get("pos_subtitle_y", self.pos_subtitle_y.value())))
        self.pos_tiny_note_y.setValue(float(data.get("pos_tiny_note_y", self.pos_tiny_note_y.value())))
        self.pos_info_x.setValue(float(data.get("pos_info_x", self.pos_info_x.value())))
        self.pos_info_y.setValue(float(data.get("pos_info_y", self.pos_info_y.value())))
        self.pos_top_right_x.setValue(float(data.get("pos_top_right_x", self.pos_top_right_x.value())))
        self.pos_top_right_y.setValue(float(data.get("pos_top_right_y", self.pos_top_right_y.value())))
        self.pos_bottom_center_y.setValue(float(data.get("pos_bottom_center_y", self.pos_bottom_center_y.value())))

        self.img_logo_path = data.get("img_logo_path", self.img_logo_path)
        self.img_icon1_path = data.get("img_icon1_path", self.img_icon1_path)
        self.img_icon2_path = data.get("img_icon2_path", self.img_icon2_path)
        self.img_icon3_path = data.get("img_icon3_path", self.img_icon3_path)

        self.lbl_logo.setText(f"Logo: {self.img_logo_path}")
        self.lbl_i1.setText(f"Icon 1: {self.img_icon1_path}")
        self.lbl_i2.setText(f"Icon 2: {self.img_icon2_path}")
        self.lbl_i3.setText(f"Icon 3: {self.img_icon3_path}")

        wanted = data.get("selected_store_id", None)
        if wanted is not None:
            idx = self.store_combo.findData(wanted)
            if idx >= 0:
                self.store_combo.setCurrentIndex(idx)

        self.excel_export_path = data.get("excel_export_path", self.excel_export_path)

    def load_settings_from_disk(self):
        if not os.path.exists(self.SETTINGS_FILE):
            return
        try:
            with open(self.SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.apply_settings(data)
        except Exception as e:
            print("Failed to load settings:", e)

    def save_settings_to_disk(self):
        try:
            data = self.collect_settings()
            with open(self.SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print("Failed to save settings:", e)

    def closeEvent(self, event):
        self.save_settings_to_disk()
        super().closeEvent(event)

    # -------- preview --------

    def update_preview(self):
        try:
            text_scale = self.s_text.value() / 100.0
            logo_scale = self.s_logo.value() / 100.0
            qr_scale = self.s_qr.value() / 100.0
            icon_scale = self.s_icon.value() / 100.0

            img = render_label(
                size=720,
                subtitle=self.f_subtitle.text().strip(),
                tiny_note=self.f_tiny_note.text().strip(),
                serial_value=self.f_serial.text().strip(),
                mac_value=self.f_mac.text().strip(),
                user_value=self.f_user.text().strip(),
                password_value=self.f_password.text().strip(),
                input_value=self.f_input.text().strip(),
                top_right_text=self.f_top_right.text().strip(),
                bottom_center_text=self.f_bottom_center.text().strip(),
                qr1_data=self.f_qr1.text().strip(),
                qr2_data=self.f_qr2.text().strip(),
                qr1_caption=self.f_qr1_cap.text().strip(),
                qr2_caption=self.f_qr2_cap.text().strip(),
                img_logo_path=self.img_logo_path,
                img_icon1_path=self.img_icon1_path,
                img_icon2_path=self.img_icon2_path,
                img_icon3_path=self.img_icon3_path,
                text_scale=text_scale,
                logo_scale=logo_scale,
                qr_scale=qr_scale,
                icon_scale=icon_scale,

                logo_x=float(self.pos_logo_x.value()),
                logo_y=float(self.pos_logo_y.value()),

                subtitle_y=float(self.pos_subtitle_y.value()),
                tiny_note_y=float(self.pos_tiny_note_y.value()),
                info_x=float(self.pos_info_x.value()),
                info_y=float(self.pos_info_y.value()),
                top_right_x=float(self.pos_top_right_x.value()),
                top_right_y=float(self.pos_top_right_y.value()),
                bottom_center_y=float(self.pos_bottom_center_y.value()),
            )
            self.label_preview.setPixmap(label_to_qpix(img, PREVIEW_SIZE_PX))
            self.save_settings_to_disk()
        except Exception as e:
            self.label_preview.setText(f"Preview error:\n{e}")

    # -------- images --------

    def pick_image(self, key: str):
        path, _ = QFileDialog.getOpenFileName(self, "Choose PNG", "", "PNG Images (*.png)")
        if not path:
            return
        if key == "logo":
            self.img_logo_path = path
            self.lbl_logo.setText(f"Logo: {path}")
        elif key == "i1":
            self.img_icon1_path = path
            self.lbl_i1.setText(f"Icon 1: {path}")
        elif key == "i2":
            self.img_icon2_path = path
            self.lbl_i2.setText(f"Icon 2: {path}")
        elif key == "i3":
            self.img_icon3_path = path
            self.lbl_i3.setText(f"Icon 3: {path}")
        self.update_preview()

    # -------- ports --------

    def refresh_ports(self):
        self.port_combo.clear()
        ports = list_com_port_infos()
        if not ports:
            self.port_combo.addItem("No COM ports found", None)
            return
        for p in ports:
            desc = getattr(p, "description", "") or ""
            label = f"{p.device} — {desc}" if desc else p.device
            self.port_combo.addItem(label, p.device)

    def on_auto_detect(self):
        detected = auto_detect_esp32_port()
        if not detected:
            return
        idx = self.port_combo.findData(detected)
        if idx >= 0:
            self.port_combo.setCurrentIndex(idx)
        mac = read_mac_from_port_fast(detected)
        if mac and MAC_RE.match(mac):
            self.f_mac.setText(mac)
            self.update_preview()

    def on_read_mac(self):
        port = self.port_combo.currentData()
        if not port:
            QMessageBox.warning(self, "No port", "Select a valid COM port.")
            return
        mac = read_mac_from_port_fast(port)
        if not mac:
            QMessageBox.warning(self, "Not found",
                                "Could not read MAC.\n\nTips:\n- Close Arduino Serial Monitor\n- Baud=115200\n- Use data USB cable")
            return
        if not MAC_RE.match(mac):
            QMessageBox.warning(self, "Invalid MAC", f"Read MAC but format looks wrong:\n{mac}")
            return
        self.f_mac.setText(mac)
        self.update_preview()

    # -------- stores --------

    def refresh_stores(self):
        self.store_combo.clear()
        self.store_combo.addItem("— Select store —", None)
        for sid, name in list_stores():
            self.store_combo.addItem(name, sid)

    def on_add_store(self):
        name = self.new_store.text().strip()
        if not name:
            QMessageBox.warning(self, "Missing store name", "Enter a store name.")
            return
        try:
            create_store(name)
            self.new_store.clear()
            self.refresh_stores()
        except Exception as e:
            QMessageBox.critical(self, "Add store failed", str(e))

    # -------- save --------

    def on_save(self):
        mac = self.f_mac.text().strip().upper()
        if not MAC_RE.match(mac):
            QMessageBox.warning(self, "Invalid MAC", "MAC must be AA:BB:CC:DD:EE:FF")
            return
        try:
            serial, admin_password = register_device(MODEL, mac)

            store_id = self.store_combo.currentData()
            if store_id is not None:
                assign_device_to_store(serial, int(store_id))

            self.f_serial.setText(serial)
            self.f_password.setText(admin_password)

            self.status1.setText(f"Serial: {serial} | MAC: {mac}")
            self.status2.setText(f"Admin password: {admin_password}")

            self.update_preview()
        except Exception as e:
            QMessageBox.critical(self, "Save failed", str(e))

    # -------- export Excel --------

    def on_export_excel(self):
        if not self.excel_export_path:
            default_path = os.path.abspath("devices_export.xlsx")
            path, _ = QFileDialog.getSaveFileName(
                self,
                "Choose Excel file location",
                default_path,
                "Excel Files (*.xlsx)"
            )
            if not path:
                return
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self.excel_export_path = path
            self.save_settings_to_disk()

        try:
            export_devices_to_excel_pro(self.excel_export_path)
            QMessageBox.information(self, "Export updated", f"Excel updated:\n{self.excel_export_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    # -------- export label --------

    def on_export_label_png(self):
        default_name = f"label_{self.f_serial.text().strip() or 'VR01'}.png"
        default_path = os.path.abspath(default_name)
        path, _ = QFileDialog.getSaveFileName(self, "Save Label PNG", default_path, "PNG Images (*.png)")
        if not path:
            return
        try:
            text_scale = self.s_text.value() / 100.0
            logo_scale = self.s_logo.value() / 100.0
            qr_scale = self.s_qr.value() / 100.0
            icon_scale = self.s_icon.value() / 100.0

            img = render_label(
                size=LABEL_EXPORT_SIZE,
                subtitle=self.f_subtitle.text().strip(),
                tiny_note=self.f_tiny_note.text().strip(),
                serial_value=self.f_serial.text().strip(),
                mac_value=self.f_mac.text().strip(),
                user_value=self.f_user.text().strip(),
                password_value=self.f_password.text().strip(),
                input_value=self.f_input.text().strip(),
                top_right_text=self.f_top_right.text().strip(),
                bottom_center_text=self.f_bottom_center.text().strip(),
                qr1_data=self.f_qr1.text().strip(),
                qr2_data=self.f_qr2.text().strip(),
                qr1_caption=self.f_qr1_cap.text().strip(),
                qr2_caption=self.f_qr2_cap.text().strip(),
                img_logo_path=self.img_logo_path,
                img_icon1_path=self.img_icon1_path,
                img_icon2_path=self.img_icon2_path,
                img_icon3_path=self.img_icon3_path,
                text_scale=text_scale,
                logo_scale=logo_scale,
                qr_scale=qr_scale,
                icon_scale=icon_scale,

                logo_x=float(self.pos_logo_x.value()),
                logo_y=float(self.pos_logo_y.value()),

                subtitle_y=float(self.pos_subtitle_y.value()),
                tiny_note_y=float(self.pos_tiny_note_y.value()),
                info_x=float(self.pos_info_x.value()),
                info_y=float(self.pos_info_y.value()),
                top_right_x=float(self.pos_top_right_x.value()),
                top_right_y=float(self.pos_top_right_y.value()),
                bottom_center_y=float(self.pos_bottom_center_y.value()),
            )
            img.save(path, format="PNG")
            QMessageBox.information(self, "Saved", f"Label saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export label failed", str(e))

    # -------- database admin actions --------

    def on_clear_devices_only(self):
        reply = QMessageBox.question(
            self,
            "Clear devices data",
            "Are you sure you want to erase ALL DEVICES data?\n\n"
            "- Devices table will be cleared\n"
            "- Assignments will be cleared\n"
            "- Stores will NOT be deleted\n\n"
            "This action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        reply2 = QMessageBox.question(
            self,
            "Confirm erase",
            "Final confirmation:\nErase ALL devices data now?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply2 != QMessageBox.Yes:
            return

        try:
            clear_devices_only()
            self.status1.setText("Serial: —")
            self.status2.setText("Admin password: —")
            QMessageBox.information(self, "Cleared", "All devices data has been erased.\nStores are kept.")
        except Exception as e:
            QMessageBox.critical(self, "Clear failed", str(e))

    def on_delete_one_device(self):
        serial_value, ok = QInputDialog.getText(
            self,
            "Delete one device",
            "Enter the device Serial to delete (example: VR01-2603-0001):"
        )
        if not ok:
            return

        serial_value = (serial_value or "").strip()
        if not serial_value:
            QMessageBox.warning(self, "Missing serial", "Please enter a serial number.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm delete",
            f"Are you sure you want to delete this device?\n\n{serial_value}\n\nThis cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        try:
            deleted = delete_one_device_by_serial(serial_value)
            if deleted <= 0:
                QMessageBox.information(self, "Not found", "No device was deleted (serial not found).")
            else:
                QMessageBox.information(self, "Deleted", f"Device deleted:\n{serial_value}")
        except Exception as e:
            QMessageBox.critical(self, "Delete failed", str(e))

    def on_delete_one_store(self):
        store_id = self.store_combo.currentData()
        store_name = self.store_combo.currentText()

        if store_id is None:
            QMessageBox.warning(self, "No store selected", "Select a store first, then click Delete one store…")
            return

        reply = QMessageBox.question(
            self,
            "Confirm delete store",
            f"Are you sure you want to delete this store?\n\n{store_name}\n\n"
            "Devices will NOT be deleted.\nAssignments to this store will be removed.\n\n"
            "This cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        try:
            deleted = delete_one_store_by_id(int(store_id))
            if deleted <= 0:
                QMessageBox.information(self, "Not deleted", "Store was not deleted (not found or schema mismatch).")
            else:
                QMessageBox.information(self, "Deleted", f"Store deleted:\n{store_name}")
            self.refresh_stores()
        except Exception as e:
            QMessageBox.critical(self, "Delete failed", str(e))


if __name__ == "__main__":
    # Ensure DB exists (db.py should create tables against its own db path or use default)
    init_db()
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())